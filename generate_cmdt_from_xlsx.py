import os
import re
from openpyxl import load_workbook
from xml.sax.saxutils import escape

def safe_devname(s: str) -> str:
    s = (s or "").strip()
    if not s:
        raise ValueError("DeveloperName is required.")
    if not re.match(r"^[A-Za-z][A-Za-z0-9_]*$", s):
        raise ValueError(f"Invalid DeveloperName '{s}'. Use letters/numbers/underscore; must start with a letter.")
    return s

def write_cmdt_record(out_dir, type_api, devname, label, fields):
    filename = f"{type_api}.{devname}.md-meta.xml"
    path = os.path.join(out_dir, filename)

    values_xml = []
    for field_api, value in fields.items():
        if value is None:
            continue
        value = str(value).strip()
        if value == "":
            continue
        values_xml.append(
            f"""  <values>
    <field>{escape(field_api)}</field>
    <value xsi:type="xsd:string">{escape(value)}</value>
  </values>"""
        )

    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<CustomMetadata xmlns="http://soap.sforce.com/2006/04/metadata"
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xmlns:xsd="http://www.w3.org/2001/XMLSchema-instance">
  <label>{escape(label)}</label>
{os.linesep.join(values_xml)}
</CustomMetadata>
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(xml)

def main():
    xlsx_path = "Trigger_Risk_Rules_Pack_v1.xlsx"   # <-- keep name same as the downloaded file
    out_dir = "force-app/main/default/customMetadata"
    type_api = "Trigger_Risk_Rule__mdt"

    os.makedirs(out_dir, exist_ok=True)

    wb = load_workbook(xlsx_path)
    ws = wb["Rules"]

    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    required = ["DeveloperName*", "MasterLabel*", "Severity__c* (High/Medium/Low)"]
    for r in required:
        if r not in idx:
            raise ValueError(f"Missing required column: {r}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        devname_raw = row[idx["DeveloperName*"]]
        if devname_raw is None or str(devname_raw).strip() == "":
            continue

        devname = safe_devname(str(devname_raw))
        label = str(row[idx["MasterLabel*"]] or devname).strip()

        fields = {
            "Rule_Key__c": row[idx.get("Rule_Key__c", -1)] if idx.get("Rule_Key__c") is not None else None,
            "Severity__c": row[idx["Severity__c* (High/Medium/Low)"]],
            "Category__c": row[idx.get("Category__c", -1)] if idx.get("Category__c") is not None else None,
            "Pattern__c": row[idx.get("Pattern__c (blank if engine-driven)", -1)] if idx.get("Pattern__c (blank if engine-driven)") is not None else None,
            "Message__c": row[idx.get("Message__c", -1)] if idx.get("Message__c") is not None else None,
        }

        write_cmdt_record(out_dir, type_api, devname, label, fields)

    print(f"Done. Wrote CMDT files to: {out_dir}")

if __name__ == "__main__":
    main()
